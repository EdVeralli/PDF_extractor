import pdfplumber
import re
import os
import csv
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def duration_to_seconds(duration_str):
    """
    Convierte una duración en formato "XdXhXmXs" a segundos.
    
    Args:
        duration_str: String con formato como "06d 23h 59m 41s" o "00s"
    
    Returns:
        int: Total de segundos
    """
    if not duration_str:
        return 0
    
    # Patrones para extraer días, horas, minutos y segundos
    patterns = {
        'days': r'(\d+)d',
        'hours': r'(\d+)h',
        'minutes': r'(\d+)m',
        'seconds': r'(\d+)s'
    }
    
    total_seconds = 0
    
    # Extraer cada componente
    for unit, pattern in patterns.items():
        match = re.search(pattern, duration_str)
        if match:
            value = int(match.group(1))
            if unit == 'days':
                total_seconds += value * 86400  # 24 * 60 * 60
            elif unit == 'hours':
                total_seconds += value * 3600   # 60 * 60
            elif unit == 'minutes':
                total_seconds += value * 60
            elif unit == 'seconds':
                total_seconds += value
    
    return total_seconds

def seconds_to_duration_format(total_seconds):
    """
    Convierte segundos totales a formato "X días, Y horas, Z minutos, W segundos"
    
    Args:
        total_seconds: int, total de segundos
    
    Returns:
        str: Duración en formato legible
    """
    if total_seconds == 0:
        return "0 segundos"
    
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    
    parts = []
    if days > 0:
        parts.append(f"{days} día{'s' if days != 1 else ''}")
    if hours > 0:
        parts.append(f"{hours} hora{'s' if hours != 1 else ''}")
    if minutes > 0:
        parts.append(f"{minutes} minuto{'s' if minutes != 1 else ''}")
    if seconds > 0:
        parts.append(f"{seconds} segundo{'s' if seconds != 1 else ''}")
    
    return ", ".join(parts)

def create_summary_excel(main_url_csv_path, excel_output_path, data_path):
    """
    Crea un archivo Excel con el resumen del downtime total.
    
    Args:
        main_url_csv_path: Ruta al archivo CSV filtrado
        excel_output_path: Ruta de salida para el archivo Excel
        data_path: Ruta a la carpeta de datos para calcular estadísticas
    """
    # Calcular las estadísticas para obtener el porcentaje y los totales
    stats = calculate_downtime_statistics(data_path)
    
    # El downtime total es la suma de los downtimes de main_url y ash_url
    total_downtime_seconds = stats['main_url_downtime'] + stats['ash_url_downtime']
    
    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Downtime"
    
    # Configurar el ancho de las columnas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    
    # Definir estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado
    ws['A1'] = "RESUMEN DE DOWNTIME"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Fila 2 - Porcentaje general
    ws['A2'] = "Uptime/Downtime (% General)"
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].border = cell_border
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['B2'] = f"{stats['ash_downtime_percentage']:.4f}%"
    ws['B2'].border = cell_border
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].font = Font(bold=True)
    
    # Fila 3 - Downtime total en formato hh:mm:ss
    ws['A3'] = "Uptime/Downtime (Home+ASH hh:mm:ss)"
    ws['A3'].fill = header_fill
    ws['A3'].font = header_font
    ws['A3'].border = cell_border
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Convertir segundos a formato hh:mm:ss
    hours = total_downtime_seconds // 3600
    minutes = (total_downtime_seconds % 3600) // 60
    seconds = total_downtime_seconds % 60
    time_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    ws['B3'] = time_formatted
    ws['B3'].border = cell_border
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B3'].font = Font(bold=True)
    
    # Información adicional
    ws['A5'] = "Información adicional:"
    ws['A5'].font = Font(bold=True, size=11)
    
    ws['A6'] = "URLs analizadas:"
    ws['B6'] = "Home (main) + ASH"
    
    ws['A7'] = "Downtime Home (segundos):"
    ws['B7'] = str(stats['main_url_downtime'])
    
    ws['A8'] = "Downtime ASH (segundos):"
    ws['B8'] = str(stats['ash_url_downtime'])
    
    ws['A9'] = "Total downtime (segundos):"
    ws['B9'] = str(total_downtime_seconds)
    
    ws['A10'] = "Tiempo total sistema (segundos):"
    ws['B10'] = str(stats['total_time'])
    
    ws['A11'] = "Fecha de generación:"
    ws['B11'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Aplicar bordes a las celdas de información adicional
    for row in range(6, 12):
        ws[f'A{row}'].border = cell_border
        ws[f'B{row}'].border = cell_border
    
    # Ajustar altura de las filas principales
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25
    
    # Guardar el archivo
    wb.save(excel_output_path)
    
    return total_downtime_seconds, time_formatted

def extract_uptime_stats(pdf_path, target_url="https://buenosaires.gob.ar/"):
    """
    Extrae los datos de Uptime Stats para una URL específica del PDF de PRTG.
    
    Args:
        pdf_path: Ruta al archivo PDF
        target_url: URL objetivo (por defecto: https://buenosaires.gob.ar/)
    
    Returns:
        dict: Diccionario con los datos de uptime o None si no se encuentra
    """
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            
            # Buscar si esta página contiene la URL objetivo
            if target_url in text:
                lines = text.split('\n')
                
                for i, line in enumerate(lines):
                    # Buscar la línea que contiene el Probe, Group, Device con nuestra URL exacta
                    if "Probe, Group, Device:" in line and target_url in line:
                        # Verificar que la URL no tiene subdominios adicionales
                        if "/tramites" not in line and "/educacion" not in line and line.strip().endswith(target_url):
                            # Buscar Uptime Stats en las siguientes líneas
                            for j in range(i, min(i + 10, len(lines))):
                                if "Uptime Stats:" in lines[j]:
                                    uptime_line = lines[j]
                                    
                                    # Extraer los valores usando regex
                                    up_pattern = r'Up:\s*([\d\.]+)\s*(%)\s*\[([^\]]+)\]'
                                    down_pattern = r'Down:\s*([\d\.]+)\s*(%)\s*\[([^\]]+)\]'
                                    
                                    up_match = re.search(up_pattern, uptime_line)
                                    down_match = re.search(down_pattern, uptime_line)
                                    
                                    if up_match and down_match:
                                        result = {
                                            'archivo_pdf': os.path.basename(pdf_path),
                                            'url': target_url,
                                            'uptime_porcentaje': up_match.group(1),
                                            'uptime_unidad': up_match.group(2),
                                            'uptime_duracion': up_match.group(3),
                                            'downtime_porcentaje': down_match.group(1),
                                            'downtime_unidad': down_match.group(2),
                                            'downtime_duracion': down_match.group(3),
                                            'pagina': page_num + 1
                                        }
                                        
                                        # Buscar información adicional
                                        for k in range(i-5, min(i + 15, len(lines))):
                                            if k >= 0:
                                                if "Report Time Span:" in lines[k]:
                                                    result['periodo_reporte'] = lines[k].split("Report Time Span:")[1].strip()
                                                elif "Report Hours:" in lines[k]:
                                                    result['horas_reporte'] = lines[k].split("Report Hours:")[1].strip()
                                                elif "Sensor Type:" in lines[k]:
                                                    result['tipo_sensor'] = lines[k].split("Sensor Type:")[1].strip()
                                                elif "Average (Loading time):" in lines[k] or "Average (Loading Time):" in lines[k]:
                                                    result['tiempo_carga_promedio'] = lines[k].split(":")[-1].strip()
                                        
                                        return result
    
    return None

def extract_all_urls_from_pdf(pdf_path):
    """
    Extrae todos los datos de uptime de todas las URLs en un PDF.
    
    Args:
        pdf_path: Ruta al archivo PDF
    
    Returns:
        list: Lista de diccionarios con los datos de uptime de cada URL
    """
    results = []
    processed_urls = set()
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            lines = text.split('\n')
            
            current_url = None
            current_probe_line = None
            
            for i, line in enumerate(lines):
                # Detectar línea de Probe, Group, Device
                if "Probe, Group, Device:" in line:
                    # Extraer la URL de diferentes formatos posibles
                    if "buenosaires.gob.ar" in line:
                        # Manejar diferentes formatos de URL en la línea
                        if "> https://buenosaires.gob.ar/" in line:
                            current_url = "https://buenosaires.gob.ar/"
                        elif "buenosaires.gob.ar/tramites" in line:
                            current_url = "buenosaires.gob.ar/tramites"
                        elif "buenosaires.gob.ar/educacion" in line:
                            current_url = "buenosaires.gob.ar/educacion"
                        elif "nba-drupal.buenosaires.gob.ar" in line:
                            current_url = "nba-drupal.buenosaires.gob.ar"
                        elif "ash.buenosaires.gob.ar" in line:
                            current_url = "ash.buenosaires.gob.ar/"
                        
                        current_probe_line = line
                
                # Buscar Uptime Stats
                elif "Uptime Stats:" in line and current_url and current_url not in processed_urls:
                    up_pattern = r'Up:\s*([\d\.]+)\s*(%)\s*\[([^\]]+)\]'
                    down_pattern = r'Down:\s*([\d\.]+)\s*(%)\s*\[([^\]]+)\]'
                    
                    up_match = re.search(up_pattern, line)
                    down_match = re.search(down_pattern, line)
                    
                    if up_match and down_match:
                        result = {
                            'archivo_pdf': os.path.basename(pdf_path),
                            'url': current_url,
                            'uptime_porcentaje': up_match.group(1),
                            'uptime_unidad': up_match.group(2),
                            'uptime_duracion': up_match.group(3),
                            'downtime_porcentaje': down_match.group(1),
                            'downtime_unidad': down_match.group(2),
                            'downtime_duracion': down_match.group(3),
                            'pagina': page_num + 1
                        }
                        
                        # Buscar información adicional
                        for k in range(max(0, i-10), min(i + 10, len(lines))):
                            if "Report Time Span:" in lines[k]:
                                result['periodo_reporte'] = lines[k].split("Report Time Span:")[1].strip()
                            elif "Report Hours:" in lines[k]:
                                result['horas_reporte'] = lines[k].split("Report Hours:")[1].strip()
                            elif "Sensor Type:" in lines[k]:
                                result['tipo_sensor'] = lines[k].split("Sensor Type:")[1].strip()
                            elif "Average (Loading time):" in lines[k] or "Average (Loading Time):" in lines[k]:
                                result['tiempo_carga_promedio'] = lines[k].split(":")[-1].strip()
                        
                        results.append(result)
                        processed_urls.add(current_url)
    
    return results

def calculate_downtime_statistics(data_path):
    """
    Calcula las estadísticas de uptime y downtime para los CSV filtrados.
    
    Args:
        data_path: Ruta a la carpeta de datos
        
    Returns:
        dict: Diccionario con las estadísticas calculadas
    """
    stats = {
        'main_url_uptime': 0,
        'main_url_downtime': 0,
        'ash_url_uptime': 0,
        'ash_url_downtime': 0,
        'total_time': 0,
        'ash_downtime_percentage': 0
    }
    
    # Leer uptime_report_main_url.csv
    main_csv_path = data_path / "uptime_report_main_url.csv"
    if main_csv_path.exists():
        with open(main_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                if 'uptime_segundos' in row and row['uptime_segundos']:
                    try:
                        stats['main_url_uptime'] += int(row['uptime_segundos'])
                    except ValueError:
                        pass
                if 'downtime_segundos' in row and row['downtime_segundos']:
                    try:
                        stats['main_url_downtime'] += int(row['downtime_segundos'])
                    except ValueError:
                        pass
    
    # Leer uptime_report_ash_url.csv
    ash_csv_path = data_path / "uptime_report_ash_url.csv"
    if ash_csv_path.exists():
        with open(ash_csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            for row in reader:
                if 'uptime_segundos' in row and row['uptime_segundos']:
                    try:
                        stats['ash_url_uptime'] += int(row['uptime_segundos'])
                    except ValueError:
                        pass
                if 'downtime_segundos' in row and row['downtime_segundos']:
                    try:
                        stats['ash_url_downtime'] += int(row['downtime_segundos'])
                    except ValueError:
                        pass
    
    # Calcular el tiempo total (uptime + downtime de ambos CSV)
    stats['total_time'] = (stats['main_url_uptime'] + stats['main_url_downtime'] + 
                          stats['ash_url_uptime'] + stats['ash_url_downtime'])
    
    # Calcular el porcentaje del downtime de ASH respecto al tiempo total
    if stats['total_time'] > 0:
        stats['ash_downtime_percentage'] = (stats['ash_url_downtime'] / stats['total_time']) * 100
    
    return stats

def process_all_pdfs(data_folder="../data", output_file="../data/uptime_report.csv"):
    """
    Procesa todos los PDFs en la carpeta data y genera un archivo CSV con los resultados.
    
    Args:
        data_folder: Ruta a la carpeta con los PDFs (relativa desde src)
        output_file: Ruta del archivo CSV de salida
    """
    # Obtener la ruta absoluta de la carpeta src
    src_path = Path(__file__).parent.absolute()
    
    # Construir la ruta absoluta a la carpeta data
    data_path = src_path.parent / "data"
    
    # Verificar que la carpeta existe
    if not data_path.exists():
        print(f"Error: La carpeta {data_path} no existe")
        return
    
    # Buscar todos los archivos PDF en la carpeta
    pdf_files = list(data_path.glob("*.pdf"))
    
    if not pdf_files:
        print(f"No se encontraron archivos PDF en {data_path}")
        return
    
    print(f"Encontrados {len(pdf_files)} archivos PDF para procesar:")
    for pdf in pdf_files:
        print(f"  - {pdf.name}")
    
    all_results = []
    
    # Procesar cada PDF
    for pdf_file in pdf_files:
        print(f"\nProcesando: {pdf_file.name}")
        try:
            # Primero intentar extraer específicamente https://buenosaires.gob.ar/
            result = extract_uptime_stats(str(pdf_file))
            if result:
                all_results.append(result)
                print(f"  ✓ Encontrado: {result['url']}")
            
            # Luego extraer todas las otras URLs del PDF
            other_results = extract_all_urls_from_pdf(str(pdf_file))
            for res in other_results:
                if res['url'] != "https://buenosaires.gob.ar/":  # Evitar duplicados
                    all_results.append(res)
                    print(f"  ✓ Encontrado: {res['url']}")
                    
        except Exception as e:
            print(f"  ✗ Error procesando {pdf_file.name}: {str(e)}")
    
    # Generar el archivo CSV
    if all_results:
        # Definir las columnas en el orden deseado
        fieldnames = [
            'archivo_pdf',
            'url',
            'uptime_porcentaje',
            'uptime_unidad',
            'uptime_duracion',
            'uptime_segundos',
            'downtime_porcentaje',
            'downtime_unidad',
            'downtime_duracion',
            'downtime_segundos',
            'periodo_reporte',
            'horas_reporte',
            'tipo_sensor',
            'tiempo_carga_promedio',
            'pagina'
        ]
        
        # Convertir duraciones a segundos para todos los resultados
        for result in all_results:
            result['uptime_segundos'] = duration_to_seconds(result.get('uptime_duracion', ''))
            result['downtime_segundos'] = duration_to_seconds(result.get('downtime_duracion', ''))
        
        # Escribir el archivo CSV principal con TODOS los resultados
        output_path = data_path / "uptime_report.csv"
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
            
            # Escribir encabezados
            writer.writeheader()
            
            # Escribir datos
            for result in all_results:
                # Asegurar que todos los campos existan
                for field in fieldnames:
                    if field not in result:
                        result[field] = ''
                writer.writerow(result)
        
        print(f"\n✅ Archivo CSV generado exitosamente: {output_path}")
        print(f"   Total de registros: {len(all_results)}")
        
        # Filtrar solo los resultados de https://buenosaires.gob.ar/
        main_url_results = [r for r in all_results if r['url'] == "https://buenosaires.gob.ar/"]
        
        if main_url_results:
            # Escribir el archivo CSV filtrado solo con https://buenosaires.gob.ar/
            filtered_output_path = data_path / "uptime_report_main_url.csv"
            with open(filtered_output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
                
                # Escribir encabezados
                writer.writeheader()
                
                # Escribir datos filtrados
                for result in main_url_results:
                    writer.writerow(result)
            
            print(f"\n✅ Archivo CSV filtrado generado exitosamente: {filtered_output_path}")
            print(f"   Total de registros para https://buenosaires.gob.ar/: {len(main_url_results)}")
            
            # Calcular las estadísticas antes de crear el Excel
            stats = calculate_downtime_statistics(data_path)
            
            # Crear el archivo Excel de resumen
            excel_output_path = data_path / "resumen_downtime.xlsx"
            total_seconds, formatted_duration = create_summary_excel(filtered_output_path, excel_output_path, data_path)
            
            print(f"\n✅ Archivo Excel de resumen generado exitosamente: {excel_output_path}")
            print(f"   Downtime total (Home + ASH): {formatted_duration}")
            print(f"   Porcentaje downtime ASH: {stats['ash_downtime_percentage']:.4f}%")
        
        # Filtrar solo los resultados de ash.buenosaires.gob.ar/
        ash_url_results = [r for r in all_results if r['url'] == "ash.buenosaires.gob.ar/"]
        
        if ash_url_results:
            # Escribir el archivo CSV filtrado solo con ash.buenosaires.gob.ar/
            ash_output_path = data_path / "uptime_report_ash_url.csv"
            with open(ash_output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
                
                # Escribir encabezados
                writer.writeheader()
                
                # Escribir datos filtrados
                for result in ash_url_results:
                    writer.writerow(result)
            
            print(f"\n✅ Archivo CSV filtrado para ASH generado exitosamente: {ash_output_path}")
            print(f"   Total de registros para ash.buenosaires.gob.ar/: {len(ash_url_results)}")
        
        # Calcular y mostrar las estadísticas de downtime
        if main_url_results or ash_url_results:
            print("\n" + "="*60)
            print("ANÁLISIS DE ESTADÍSTICAS DE UPTIME/DOWNTIME")
            print("="*60)
            
            stats = calculate_downtime_statistics(data_path)
            
            print(f"\n📊 Resumen de tiempos (en segundos):")
            print(f"   └─ uptime_report_main_url.csv:")
            print(f"      • Uptime:   {stats['main_url_uptime']:,} segundos")
            print(f"      • Downtime: {stats['main_url_downtime']:,} segundos")
            
            print(f"\n   └─ uptime_report_ash_url.csv:")
            print(f"      • Uptime:   {stats['ash_url_uptime']:,} segundos")
            print(f"      • Downtime: {stats['ash_url_downtime']:,} segundos")
            
            print(f"\n📈 Totales acumulados:")
            print(f"   • Total Uptime (ambos CSV):   {stats['main_url_uptime'] + stats['ash_url_uptime']:,} segundos")
            print(f"   • Total Downtime (ambos CSV): {stats['main_url_downtime'] + stats['ash_url_downtime']:,} segundos")
            print(f"   • TIEMPO TOTAL:               {stats['total_time']:,} segundos")
            
            print(f"\n🎯 Análisis de porcentaje:")
            print(f"   • Downtime de ASH: {stats['ash_url_downtime']:,} segundos")
            print(f"   • Tiempo total:    {stats['total_time']:,} segundos")
            print(f"   • PORCENTAJE:      {stats['ash_downtime_percentage']:.4f}%")
            
            # Convertir a formato legible
            if stats['ash_url_downtime'] > 0:
                ash_downtime_formatted = seconds_to_duration_format(stats['ash_url_downtime'])
                total_time_formatted = seconds_to_duration_format(stats['total_time'])
                print(f"\n   En formato legible:")
                print(f"   • Downtime de ASH: {ash_downtime_formatted}")
                print(f"   • Tiempo total:    {total_time_formatted}")
        
        # Mostrar resumen
        print("\nResumen de URLs procesadas:")
        urls_by_pdf = {}
        for result in all_results:
            pdf_name = result['archivo_pdf']
            if pdf_name not in urls_by_pdf:
                urls_by_pdf[pdf_name] = []
            urls_by_pdf[pdf_name].append(result['url'])
        
        for pdf_name, urls in urls_by_pdf.items():
            print(f"\n  {pdf_name}:")
            for url in urls:
                print(f"    - {url}")
    else:
        print("\n⚠️ No se encontraron datos de uptime en ningún PDF")

def main():
    """
    Función principal del programa.
    
    Requiere instalar:
    pip install pdfplumber openpyxl
    """
    print("=== Extractor de Uptime Stats de PDFs PRTG ===")
    print(f"Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)
    
    # Procesar todos los PDFs
    process_all_pdfs()
    
    print("\nProceso completado.")

if __name__ == "__main__":
    main()